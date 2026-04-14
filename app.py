# -*- coding: utf-8 -*-
from itsdangerous import SignatureExpired, BadSignature
from itsdangerous import URLSafeTimedSerializer
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import sqlite3
from datetime import datetime, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from werkzeug.security import generate_password_hash, check_password_hash
import qrcode
import os
import time
from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import Workbook
from fpdf import FPDF
from email.mime.base import MIMEBase
from email import encoders


scheduler = BackgroundScheduler()


app = Flask(__name__)

# PRIMERO define la clave
app.secret_key = "clave_secreta_segura"

# DESPUÉS creas el serializer
serializer = URLSafeTimedSerializer(app.secret_key)

app.permanent_session_lifetime = timedelta(minutes=30)

EMAIL = "cr3109867@gmail.com"
PASSWORD = "ksjg crnr jsvo acys"

DASHBOARD_ROUTES = {
    "farmacia": "dashboard_farmacia",
    "ferreteria": "dashboard_ferreteria",
    "mercado": "dashboard_mercado",
    "tienda_de_ropa": "dashboard_tienda_de_ropa",
}

NEGOCIOS_INFO = {
    "farmacia": {
        "nombre": "Farmacia Central Salud",
        "panel": "Farmacia",
        "descripcion": "Control de medicamentos, inventario, vencimientos y ventas.",
        "propietario": "Dra. María Fernanda Ríos",
        "telefono": "+57 300 123 4567",
        "correo": "farmacia@ventasapp.com",
        "direccion": "Calle 10 # 15-20, Centro",
        "nit": "901.456.321-0",
        "icono": "fas fa-pills",
        "color": "success",
    },
    "ferreteria": {
        "nombre": "Ferretería El Tornillo Firme",
        "panel": "Ferretería",
        "descripcion": "Gestión de materiales, stock, compras y ventas del negocio.",
        "propietario": "Carlos Andrés Méndez",
        "telefono": "+57 301 222 3344",
        "correo": "ferreteria@ventasapp.com",
        "direccion": "Carrera 18 # 22-45, Zona Industrial",
        "nit": "900.987.654-1",
        "icono": "fas fa-tools",
        "color": "warning",
    },
    "mercado": {
        "nombre": "Mercado Fresco Hogar",
        "panel": "Mercado",
        "descripcion": "Seguimiento de productos, inventario diario y ventas rápidas.",
        "propietario": "Ana Lucía Gómez",
        "telefono": "+57 302 111 5566",
        "correo": "mercado@ventasapp.com",
        "direccion": "Avenida 5 # 30-12, Barrio Comercial",
        "nit": "901.222.888-3",
        "icono": "fas fa-apple-alt",
        "color": "success",
    },
    "tienda_de_ropa": {
        "nombre": "Tienda de Ropa Urban Style",
        "panel": "Tienda de Ropa",
        "descripcion": "Panel de categorías, productos destacados y rendimiento comercial.",
        "propietario": "Laura Sofía Herrera",
        "telefono": "+57 304 777 8899",
        "correo": "ropa@ventasapp.com",
        "direccion": "Centro Comercial Plaza Moda, Local 12",
        "nit": "901.654.777-5",
        "icono": "fas fa-store",
        "color": "primary",
    },
}


# ---------------------------
# HELPERS
# ---------------------------
def get_db_connection():
    conn = sqlite3.connect("ventas_app.db")
    conn.row_factory = sqlite3.Row
    conn.text_factory = lambda b: b.decode("utf-8")
    return conn


def crear_tabla_datos_negocio():
    conn = get_db_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS datos_negocio (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER NOT NULL,
            negocio TEXT NOT NULL,
            nombre_negocio TEXT,
            responsable TEXT,
            telefono TEXT,
            correo TEXT,
            direccion TEXT,
            nit TEXT,
            UNIQUE(usuario_id, negocio)
        )
    """)
    conn.commit()
    conn.close()


def obtener_info_negocio(negocio=None):
    negocio = negocio or session.get("negocio")
    if not negocio:
        return None

    info = NEGOCIOS_INFO.get(negocio, {}).copy()

    if session.get("usuario_id"):
        crear_tabla_datos_negocio()
        conn = get_db_connection()
        fila = conn.execute("""
            SELECT nombre_negocio, responsable, telefono, correo, direccion, nit
            FROM datos_negocio
            WHERE usuario_id = ? AND negocio = ?
        """, (session["usuario_id"], negocio)).fetchone()
        conn.close()

        if fila:
            info["nombre"] = fila["nombre_negocio"] or info.get("nombre", "")
            info["propietario"] = fila["responsable"] or info.get("propietario", "")
            info["telefono"] = fila["telefono"] or info.get("telefono", "")
            info["correo"] = fila["correo"] or info.get("correo", "")
            info["direccion"] = fila["direccion"] or info.get("direccion", "")
            info["nit"] = fila["nit"] or info.get("nit", "")
            info["registrado"] = True
        else:
            info["registrado"] = False

    return info


def redirigir_dashboard_por_negocio(negocio):
    endpoint = DASHBOARD_ROUTES.get(negocio, "seleccionar_negocio")
    return redirect(url_for(endpoint))


def validar_dashboard(negocio_esperado):
    if "usuario_id" not in session:
        flash("Debes iniciar sesión para acceder al panel.", "warning")
        return redirect(url_for("login")), None

    negocio_actual = session.get("negocio")
    if not negocio_actual:
        flash("Debes seleccionar un negocio antes de entrar al dashboard.", "warning")
        return redirect(url_for("seleccionar_negocio")), None

    if negocio_actual != negocio_esperado:
        flash("Te redirigimos al panel del negocio que tienes activo.", "info")
        return redirigir_dashboard_por_negocio(negocio_actual), None

    return None, negocio_actual


@app.context_processor
def inject_global_business_context():
    negocio_actual = session.get("negocio")
    return {
        "negocio_actual": negocio_actual,
        "negocio_info": obtener_info_negocio(negocio_actual),
    }


# ---------------------------
# EMAIL
# ---------------------------
def enviar_correo(destinatario, asunto, html, remitente=EMAIL):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = str(asunto)
    msg["From"] = remitente
    msg["To"] = destinatario
    msg.attach(MIMEText(html, "html", "utf-8"))

    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(EMAIL, PASSWORD)
            server.sendmail(remitente, destinatario, msg.as_string())
    except Exception as e:
        print("Error correo:", e)


# ---------------------------
# INDEX
# ---------------------------
@app.route("/")
def index():
    return render_template("index.html")


# ---------------------------
# REGISTER
# ---------------------------
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        correo = request.form["correo"]
        contraseña = request.form["contraseña"]
        nombre = request.form["nombre"]

        contraseña_hash = generate_password_hash(contraseña)

        conn = get_db_connection()

        total = conn.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0]
        rol = "admin" if total == 0 else "usuario"

        try:
            conn.execute(
                "INSERT INTO usuarios (correo, contraseña, nombre, rol) VALUES (?, ?, ?, ?)",
                (correo, contraseña_hash, nombre, rol),
            )
            conn.commit()
        except Exception:
            flash("Correo ya registrado", "danger")
            conn.close()
            return redirect(url_for("register"))

        conn.close()
        flash("Registro exitoso", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


# ---------------------------
# LOGIN
# ---------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        correo = request.form["correo"]
        contraseña = request.form["contraseña"]

        conn = get_db_connection()
        usuario = conn.execute(
            "SELECT * FROM usuarios WHERE correo=?", (correo,)
        ).fetchone()
        conn.close()

        if usuario and check_password_hash(usuario["contraseña"], contraseña):
            session.permanent = True
            session["usuario_id"] = usuario["id"]
            session["usuario_nombre"] = usuario["nombre"]
            session["rol"] = usuario["rol"] if usuario["rol"] else "usuario"
            session["negocio"] = usuario["negocio"]

            try:
                html = render_template(
                    "emails/login_notification.html",
                    nombre=usuario["nombre"]
                )

                enviar_correo(
                    usuario["correo"],
                    "Inicio de sesión exitoso",
                    html
                )

                print("✅ Correo enviado correctamente")

            except Exception as e:
                print("❌ Error enviando correo:", e)

            flash("Bienvenido " + usuario["nombre"], "success")
            return redirect(url_for("seleccionar_negocio"))

        flash("Datos incorrectos", "danger")

    return render_template("login.html")


# ---------------------------
# FORGOT PASSWORD
# ---------------------------
@app.route("/forgot_password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        correo = request.form["correo"]

        conn = get_db_connection()
        usuario = conn.execute(
            "SELECT * FROM usuarios WHERE correo=?", (correo,)
        ).fetchone()
        conn.close()

        if usuario:
            token = serializer.dumps(correo, salt="password-reset")
            link = url_for("reset_password", token=token, _external=True)

            html = f"""
            <h2>🔐 Recuperar contraseña</h2>
            <p>Hola {usuario["nombre"]},</p>
            <p>Haz clic en el botón para cambiar tu contraseña:</p>

            <a href="{link}" style="padding:10px 20px;
            background:#27ae60;color:white;text-decoration:none;border-radius:5px;">
            Cambiar contraseña
            </a>

            <p>⚠️ Este enlace expira en 15 minutos</p>
            """

            enviar_correo(correo, "Recuperar contraseña", html)

        flash("Si el correo existe, se enviará un enlace", "info")
        return redirect(url_for("login"))

    return render_template("forgot_password.html")


# ---------------------------
# RESET PASSWORD
# ---------------------------
@app.route("/reset_password/<token>", methods=["GET", "POST"])
def reset_password(token):
    try:
        correo = serializer.loads(token, salt="password-reset", max_age=900)

    except SignatureExpired:
        flash("⏰ El enlace expiró", "danger")
        return redirect(url_for("forgot_password"))

    except BadSignature:
        flash("⚠️ Token inválido", "danger")
        return redirect(url_for("login"))

    conn = get_db_connection()
    usuario = conn.execute(
        "SELECT * FROM usuarios WHERE correo=?", (correo,)
    ).fetchone()

    if not usuario:
        conn.close()
        flash("Usuario no encontrado", "danger")
        return redirect(url_for("login"))

    if request.method == "POST":
        nueva = request.form["contraseña"]

        if len(nueva) < 6:
            flash("Mínimo 6 caracteres", "warning")
            return redirect(request.url)

        hash_nueva = generate_password_hash(nueva)

        conn.execute(
            "UPDATE usuarios SET contraseña=? WHERE id=?",
            (hash_nueva, usuario["id"])
        )
        conn.commit()
        conn.close()

        flash("✅ Contraseña actualizada", "success")
        return redirect(url_for("login"))

    conn.close()
    return render_template("reset_password.html")


# ---------------------------
# LOGOUT
# ---------------------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))


#----------------------------
#reportes programa
#----------------------------
import sqlite3
from datetime import datetime
from apscheduler.schedulers.background import BackgroundScheduler

def enviar_reportes_programados():
    conn = sqlite3.connect("ventas_app.db")
    conn.row_factory = sqlite3.Row
    reportes = conn.execute("SELECT * FROM reportes_programados").fetchall()
    conn.close()

    current_time = datetime.now().strftime("%H:%M")
    current_day = datetime.now().weekday()  # 0 = lunes, 6 = domingo
    current_day_of_month = datetime.now().day

    for r in reportes:
        # Verificar si la hora coincide
        if r["hora"] == current_time:
            if r["frecuencia"] == "diario":
                generar_y_enviar_reporte(r)
            elif r["frecuencia"] == "semanal" and current_day == 0:  # lunes
                generar_y_enviar_reporte(r)
            elif r["frecuencia"] == "mensual" and current_day_of_month == 1:  # primer día del mes
                generar_y_enviar_reporte(r)

# Iniciar el scheduler en segundo plano
scheduler = BackgroundScheduler()
scheduler.add_job(enviar_reportes_programados, "interval", minutes=1)
scheduler.start()

#----------------------------
#generar y enviar reporte
#----------------------------
def generar_y_enviar_reporte(r):
    # 👉 Decidir formato del reporte (PDF por defecto, puedes ampliar a Excel o ambos)
    formato = r["formato"] if "formato" in r.keys() else "pdf"

    msg = MIMEMultipart()
    msg["From"] = EMAIL
    msg["To"] = r["destinatario"]
    msg["Subject"] = "Reporte automático de inventario"

    body = "Adjunto el reporte automático de inventario."
    msg.attach(MIMEText(body, "plain"))

    try:
        if formato == "pdf":
            filename = f"reporte_{r['negocio']}.pdf"
            ruta_archivo = exportar_inventario_pdf(r['negocio'], filename)

            with open(ruta_archivo, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={filename}")
                msg.attach(part)

        elif formato == "excel":
            filename = f"reporte_{r['negocio']}.xlsx"
            ruta_archivo = exportar_inventario_excel(r['negocio'], filename)

            with open(ruta_archivo, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={filename}")
                msg.attach(part)

        elif formato == "ambos":
            # PDF
            filename_pdf = f"reporte_{r['negocio']}.pdf"
            ruta_pdf = exportar_inventario_pdf(r['negocio'], filename_pdf)
            with open(ruta_pdf, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={filename_pdf}")
                msg.attach(part)

            # Excel
            filename_excel = f"reporte_{r['negocio']}.xlsx"
            ruta_excel = exportar_inventario_excel(r['negocio'], filename_excel)
            with open(ruta_excel, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={filename_excel}")
                msg.attach(part)

    except Exception as e:
        print("❌ Error adjuntando archivo:", e)

    # Enviar correo
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(EMAIL, PASSWORD)
        server.send_message(msg)
        server.quit()
        print("✅ Reporte enviado correctamente a", r["destinatario"])
    except Exception as e:
        print("❌ Error enviando correo:", e)

#---------------------------
# Exportar reporte PDF 
#---------------------------
def exportar_inventario_pdf(negocio, filename="inventario.pdf"):
    # Ruta donde se guardará el archivo
    carpeta_reportes = os.path.join("static", "reportes")
    os.makedirs(carpeta_reportes, exist_ok=True)
    ruta = os.path.join(carpeta_reportes, filename)

    # 👉 Generar el PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=f"Reporte de inventario - {negocio}", ln=True, align="C")

    # Ejemplo: agregar productos desde la BD
    conn = get_db_connection()
    productos = conn.execute(
        "SELECT nombre, cantidad FROM productos WHERE negocio=?", (negocio,)
    ).fetchall()
    conn.close()

    for p in productos:
        pdf.cell(200, 10, txt=f"{p['nombre']} - Stock: {p['cantidad']}", ln=True)

    # Guardar el archivo en disco
    pdf.output(ruta)

    return ruta  

# ---------------------------
# Seleccionar negocio
# ---------------------------
@app.route("/seleccionar_negocio", methods=["GET", "POST"])
def seleccionar_negocio():
    if "usuario_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        negocio = request.form.get("negocio")
        session["negocio"] = negocio

        conn = get_db_connection()
        conn.execute(
            "UPDATE usuarios SET negocio=? WHERE id=?",
            (negocio, session["usuario_id"])
        )
        conn.commit()
        conn.close()

        flash(f"Negocio cambiado a {negocio.capitalize()}", "success")

        return render_template(
            "seleccionar_negocio.html",
            negocio_actual=negocio,
            mostrar_formulario=True
        )

    negocio_actual = session.get("negocio")
    return render_template(
        "seleccionar_negocio.html",
        negocio_actual=negocio_actual,
        mostrar_formulario=False
    )


# ---------------------------
# Guardar datos negocio
# ---------------------------
@app.route("/guardar_datos_negocio", methods=["POST"])
def guardar_datos_negocio():
    if "usuario_id" not in session:
        flash("Debes iniciar sesión primero.", "warning")
        return redirect(url_for("login"))

    negocio = session.get("negocio")
    if not negocio:
        flash("Debes seleccionar un negocio primero.", "warning")
        return redirect(url_for("seleccionar_negocio"))

    crear_tabla_datos_negocio()

    nombre_negocio = request.form.get("nombre_negocio", "").strip()
    responsable = request.form.get("responsable", "").strip()
    telefono = request.form.get("telefono", "").strip()
    correo = request.form.get("correo", "").strip()
    direccion = request.form.get("direccion", "").strip()
    nit = request.form.get("nit", "").strip()

    conn = get_db_connection()
    existente = conn.execute("""
        SELECT id FROM datos_negocio
        WHERE usuario_id = ? AND negocio = ?
    """, (session["usuario_id"], negocio)).fetchone()

    if existente:
        conn.execute("""
            UPDATE datos_negocio
            SET nombre_negocio=?, responsable=?, telefono=?, correo=?, direccion=?, nit=?
            WHERE usuario_id=? AND negocio=?
        """, (
            nombre_negocio, responsable, telefono, correo, direccion, nit,
            session["usuario_id"], negocio
        ))
    else:
        conn.execute("""
            INSERT INTO datos_negocio
            (usuario_id, negocio, nombre_negocio, responsable, telefono, correo, direccion, nit)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            session["usuario_id"], negocio, nombre_negocio, responsable,
            telefono, correo, direccion, nit
        ))

    conn.commit()
    conn.close()

    flash("Datos guardados correctamente.", "success")
    return redirigir_dashboard_por_negocio(negocio)


# ---------------------------
# DASHBOARD FARMACIA
# ---------------------------
@app.route("/dashboard_farmacia")
def dashboard_farmacia():
    respuesta, _ = validar_dashboard("farmacia")
    if respuesta:
        return respuesta

    inventario_farmacia = [
        {"nombre": "Paracetamol", "categoria": "Analgésicos", "stock": 50},
        {"nombre": "Amoxicilina", "categoria": "Antibióticos", "stock": 30},
        {"nombre": "Vitamina C", "categoria": "Vitaminas", "stock": 80},
        {"nombre": "Ibuprofeno", "categoria": "Analgésicos", "stock": 60},
    ]

    categorias = ["Analgésicos", "Antibióticos", "Vitaminas", "Otros"]
    ventas_categoria = [500, 300, 200, 100]

    proximos_vencer = [
        {"nombre": "Paracetamol", "lote": "A123", "vencimiento": "2026-05-10", "stock": 50},
        {"nombre": "Amoxicilina", "lote": "B456", "vencimiento": "2026-06-15", "stock": 30}
    ]

    total_productos = len(inventario_farmacia)
    stock_total = sum(item["stock"] for item in inventario_farmacia)

    return render_template(
        "dashboard_farmacia.html",
        categorias=categorias,
        ventas_categoria=ventas_categoria,
        proximos_vencer=proximos_vencer,
        total_productos=total_productos,
        stock_total=stock_total,
    )


# ---------------------------
# DASHBOARD FERRETERÍA
# ---------------------------
@app.route("/dashboard_ferreteria")
def dashboard_ferreteria():
    respuesta, _ = validar_dashboard("ferreteria")
    if respuesta:
        return respuesta

    meses = ["Enero", "Febrero", "Marzo", "Abril"]
    ventas = [800, 1200, 1500, 1000]
    materiales = [
        {"nombre": "Martillo", "unidad": "Unidad", "stock": 20, "precio": 15000},
        {"nombre": "Cemento", "unidad": "Kg", "stock": 100, "precio": 25000}
    ]
    bajo_stock = [m for m in materiales if m["stock"] < 30]
    stock_total = sum(m["stock"] for m in materiales)

    return render_template(
        "dashboard_ferreteria.html",
        meses=meses,
        ventas=ventas,
        materiales=materiales,
        bajo_stock=bajo_stock,
        stock_total=stock_total,
    )


# ---------------------------
# DASHBOARD MERCADO
# ---------------------------
@app.route("/dashboard_mercado")
def dashboard_mercado():
    respuesta, _ = validar_dashboard("mercado")
    if respuesta:
        return respuesta

    productos = [
        {"nombre": "Manzanas", "stock": 100, "precio": 2000},
        {"nombre": "Plátanos", "stock": 80, "precio": 1500},
        {"nombre": "Tomates", "stock": 50, "precio": 2500}
    ]
    ventas_diarias = [20, 35, 40, 25]

    return render_template(
        "dashboard_mercado.html",
        productos=productos,
        ventas_diarias=ventas_diarias
    )


# ---------------------------
# DASHBOARD TIENDA DE ROPA
# ---------------------------
@app.route("/dashboard_tienda_de_ropa")
def dashboard_tienda_de_ropa():
    respuesta, _ = validar_dashboard("tienda_de_ropa")
    if respuesta:
        return respuesta

    categorias = ["Ropa", "Electrodomésticos", "Juguetes"]
    ventas_categoria = [300, 500, 200]

    return render_template(
        "dashboard_tienda_de_ropa.html",
        categorias=categorias,
        ventas_categoria=ventas_categoria
    )


#----------------------------
#formulario negocio
#----------------------------
@app.route("/formulario_negocio/<tipo>", methods=["GET", "POST"])
def formulario_negocio(tipo):
    if request.method == "POST":
        # Capturar datos del formulario
        nombre = request.form["nombre"]
        patron = request.form["patron"]
        usuario = request.form["usuario"]
        descripcion = request.form["descripcion"]

        # Guardar en la base de datos
        conn = sqlite3.connect("ventas_app.db")
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO negocios (tipo, nombre, patron, usuario, descripcion)
            VALUES (?, ?, ?, ?, ?)
        """, (tipo, nombre, patron, usuario, descripcion))
        conn.commit()
        conn.close()

        # ✅ Guardar negocio actual en sesión
        session["negocio_actual"] = tipo
        session["negocio_nombre"] = nombre   # 👈 Guardamos el nombre real

        flash("✅ Negocio registrado correctamente", "success")
        # Redirigir al dashboard del negocio elegido
        return redirect(url_for("dashboard", tipo=tipo))

    # Renderizar el formulario
    return render_template("formulario_negocio.html", tipo=tipo)


# ---------------------------
# DASHBOARD GENERAL
# ---------------------------

@app.route("/dashboard/<tipo>")
def dashboard(tipo):
    conn = sqlite3.connect("ventas_app.db")
    cursor = conn.cursor()

    # Buscar datos del negocio
    cursor.execute("""
        SELECT nombre, patron, usuario, descripcion 
        FROM negocios 
        WHERE tipo = ?
    """, (tipo,))
    negocio = cursor.fetchone()

    # Buscar productos del negocio
    cursor.execute("""
        SELECT nombre, categoria, precio, cantidad 
        FROM productos 
        WHERE negocio = ?
    """, (tipo,))
    productos = cursor.fetchall()

    # Buscar ventas del negocio
    cursor.execute("""
        SELECT fecha, cliente, total 
        FROM ventas v
        JOIN usuarios u ON v.usuario_id = u.id
        WHERE u.negocio = ?
    """, (tipo,))
    ventas = cursor.fetchall()

    conn.close()

    if negocio:
        nombre, patron, usuario, descripcion = negocio
    else:
        nombre, patron, usuario, descripcion = None, None, None, None

    # ✅ Guardar también en sesión para usar en navbar y títulos
    session["negocio_actual"] = tipo
    session["negocio_nombre"] = nombre

    return render_template("dashboard.html",
                           tipo=tipo,
                           nombre=nombre,
                           patron=patron,
                           usuario=usuario,
                           descripcion=descripcion,
                           productos=productos,
                           ventas=ventas)

# ---------------------------
# USUARIOS (ADMIN)
# ---------------------------
@app.route("/usuarios")
def usuarios():
    if session.get("rol") != "admin":
        return redirect(url_for("index"))

    conn = get_db_connection()
    usuarios = conn.execute("SELECT id, nombre, correo, rol FROM usuarios").fetchall()
    conn.close()

    return render_template("usuarios.html", usuarios=usuarios)


# ---------------------------
# CAMBIAR ROL
# ---------------------------
@app.route("/cambiar_rol/<int:id>")
def cambiar_rol(id):
    if session.get("rol") != "admin":
        return redirect(url_for("index"))

    if id == session.get("usuario_id"):
        return redirect(url_for("usuarios"))

    conn = get_db_connection()
    usuario = conn.execute("SELECT rol FROM usuarios WHERE id=?", (id,)).fetchone()

    nuevo_rol = "admin" if usuario["rol"] == "usuario" else "usuario"

    conn.execute("UPDATE usuarios SET rol=? WHERE id=?", (nuevo_rol, id))
    conn.commit()
    conn.close()

    return redirect(url_for("usuarios"))


# ---------------------------
# INVENTARIO
# ---------------------------

@app.route("/inventario")
def inventario():
    """
    Muestra el inventario filtrado por el negocio actual.
    Incluye alertas de stock bajo, control de vencimiento, QR, gráficas,
    la lista de reportes programados y conteo por categorías.
    """
    negocio_actual = session.get("negocio")

    if not negocio_actual:
        flash("Debes seleccionar un negocio antes de ver el inventario.", "warning")
        return redirect(url_for("seleccionar_negocio"))

    conn = None

    # Capturar filtros desde la URL (GET)
    categoria = request.args.get("categoria")
    stock_bajo = request.args.get("stock_bajo")
    proximos_vencer = request.args.get("proximos_vencer")

    try:
        conn = get_db_connection()
        query = "SELECT * FROM productos WHERE negocio = ?"
        params = [negocio_actual]

        # Filtro por categoría
        if categoria:
            query += " AND categoria LIKE ?"
            params.append(f"%{categoria}%")

        productos = conn.execute(query, params).fetchall()

        # Traer reportes programados del negocio (⚠️ corregido: ordena por id si no existe creado_en)
        reportes_programados = conn.execute(
            "SELECT * FROM reportes_programados WHERE negocio = ? ORDER BY id DESC",
            (negocio_actual,)
        ).fetchall()

        # Traer historial de reportes enviados
        reportes_enviados = conn.execute(
            "SELECT * FROM reportes_enviados WHERE negocio = ? ORDER BY fecha_envio DESC",
            (negocio_actual,)
        ).fetchall()

    except Exception as e:
        flash(f"Error al cargar inventario: {str(e)}", "danger")
        productos, reportes_programados, reportes_enviados = [], [], []
    finally:
        if conn:
            conn.close()

    current_date = datetime.now().date()


    # Filtro adicional en memoria (stock bajo y próximos a vencer)
    if stock_bajo:
        productos = [p for p in productos if p["cantidad"] < 5]

    if proximos_vencer:
        filtrados = []
        for p in productos:
            if p["fecha_vencimiento"]:
                try:
                    fecha_venc = datetime.strptime(p["fecha_vencimiento"], "%Y-%m-%d").date()
                    dias_restantes = (fecha_venc - current_date).days
                    if dias_restantes < 30 and dias_restantes >= 0:
                        filtrados.append(p)
                except Exception:
                    pass
        productos = filtrados

    # ---------------------------
    # Alertas inteligentes
    # ---------------------------
    alertas_stock = [p for p in productos if p["cantidad"] < 5]
    alertas_vencimiento = []
    for p in productos:
        if p["fecha_vencimiento"]:
            try:
                fecha_venc = datetime.strptime(p["fecha_vencimiento"], "%Y-%m-%d").date()
                dias_restantes = (fecha_venc - current_date).days
                if dias_restantes < 30 and dias_restantes >= 0:
                    alertas_vencimiento.append(p)
            except Exception:
                pass

    # ---------------------------
    # Datos agregados para gráficas
    # ---------------------------
    categorias_count = {}
    stock_bajo_count = len(alertas_stock)
    proximos_vencer_count = len(alertas_vencimiento)

    for p in productos:
        # Conteo por categoría (sumando cantidades)
        cat = p["categoria"] or "Sin categoría"
        categorias_count[cat] = categorias_count.get(cat, 0) + p["cantidad"]

    # ---------------------------
    # Render final (solo uno)
    # ---------------------------
    return render_template(
        "inventario.html",
        negocio_actual=negocio_actual,
        productos=productos,
        categoria=categoria,
        stock_bajo=stock_bajo,
        proximos_vencer=proximos_vencer,
        current_date=current_date,
        alertas_stock=alertas_stock,
        alertas_vencimiento=alertas_vencimiento,
        reportes_programados=reportes_programados,
        reportes_enviados=reportes_enviados,
        categorias_count=categorias_count,
        stock_bajo_count=stock_bajo_count,
        proximos_vencer_count=proximos_vencer_count
    )




# ----------------------------
# Programar reporte automático
# ----------------------------
@app.route("/programar_reporte", methods=["POST"])
def programar_reporte():
    negocio_actual = session.get("negocio")

    if not negocio_actual:
        flash("⚠️ Debes seleccionar un negocio antes de programar un reporte.", "warning")
        return redirect(url_for("inventario"))

    # Capturar datos del formulario
    frecuencia = request.form.get("frecuencia")
    hora = request.form.get("hora")
    destinatario = request.form.get("destinatario")
    formato = request.form.get("formato", "pdf")  # por defecto PDF
    tipo_reporte = request.form.get("tipo_reporte", "completo")  # por defecto inventario completo
    cc = request.form.get("cc")  # opcional

    # Validaciones básicas
    if not destinatario or not hora:
        flash("⚠️ Debes ingresar un correo y una hora válidos.", "danger")
        return redirect(url_for("inventario"))

    # Guardar en la base de datos
    conn = get_db_connection()
    conn.execute("""
        INSERT INTO reportes_programados 
        (negocio, frecuencia, hora, destinatario, formato, tipo_reporte, cc)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (negocio_actual, frecuencia, hora, destinatario, formato, tipo_reporte, cc))
    conn.commit()
    conn.close()

    flash("✅ Reporte programado correctamente", "success")
    return redirect(url_for("inventario"))


# ----------------------------
# Eliminar reporte programado
# ----------------------------
@app.route("/eliminar_reporte/<int:id>", methods=["POST"])
def eliminar_reporte(id):
    negocio_actual = session.get("negocio")

    if not negocio_actual:
        flash("⚠️ Debes seleccionar un negocio antes de eliminar un reporte.", "warning")
        return redirect(url_for("inventario"))

    try:
        conn = get_db_connection()
        conn.execute("DELETE FROM reportes_programados WHERE id = ? AND negocio = ?", (id, negocio_actual))
        conn.commit()
        conn.close()
        flash("🗑️ Reporte eliminado correctamente", "success")
    except Exception as e:
        flash(f"Error al eliminar reporte: {str(e)}", "danger")

    return redirect(url_for("inventario"))


# ----------------------------
# Editar reporte programado
# ----------------------------
@app.route("/editar_reporte/<int:id>", methods=["POST"])
def editar_reporte(id):
    negocio_actual = session.get("negocio")

    if not negocio_actual:
        flash("⚠️ Debes seleccionar un negocio antes de editar un reporte.", "warning")
        return redirect(url_for("inventario"))

    # Capturar datos del formulario
    frecuencia = request.form.get("frecuencia")
    hora = request.form.get("hora")
    destinatario = request.form.get("destinatario")
    formato = request.form.get("formato", "pdf")
    tipo_reporte = request.form.get("tipo_reporte", "completo")
    cc = request.form.get("cc")

    # Validaciones básicas
    if not destinatario or not hora:
        flash("⚠️ Debes ingresar un correo y una hora válidos.", "danger")
        return redirect(url_for("inventario"))

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Verificar si el reporte existe
        reporte = cursor.execute(
            "SELECT * FROM reportes_programados WHERE id = ? AND negocio = ?",
            (id, negocio_actual)
        ).fetchone()

        if not reporte:
            conn.close()
            flash("❌ El reporte que intentas editar ya no existe.", "danger")
            return redirect(url_for("inventario"))

        # Actualizar el reporte
        cursor.execute("""
            UPDATE reportes_programados
            SET frecuencia = ?, hora = ?, destinatario = ?, formato = ?, tipo_reporte = ?, cc = ?
            WHERE id = ? AND negocio = ?
        """, (frecuencia, hora, destinatario, formato, tipo_reporte, cc, id, negocio_actual))

        conn.commit()
        conn.close()
        flash("✏️ Reporte actualizado correctamente", "success")

    except Exception as e:
        flash(f"Error al actualizar reporte: {str(e)}", "danger")

    return redirect(url_for("inventario"))


# ---------------------------
# REGISTRAR PRODUCTO (ADMIN)
# ---------------------------
@app.route("/registrar_producto", methods=["GET", "POST"])
def registrar_producto():
    if session.get("rol") != "admin":
        flash("No tienes permisos para registrar productos.", "danger")
        return redirect(url_for("inventario"))

    negocio_actual = session.get("negocio")
    if not negocio_actual:
        flash("Debes seleccionar un negocio antes de registrar productos.", "warning")
        return redirect(url_for("seleccionar_negocio"))

    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            categoria = request.form["categoria"].strip()
            precio = float(request.form["precio"])
            cantidad = int(request.form["cantidad"])
            proveedor = request.form.get("proveedor", "").strip()
            codigo_barras = request.form.get("codigo_barras", "").strip()
            fecha_vencimiento = request.form.get("fecha_vencimiento") or None

            qr_data = codigo_barras if codigo_barras else nombre
            qr_img = qrcode.make(qr_data)
            qr_folder = os.path.join("static", "qr")
            os.makedirs(qr_folder, exist_ok=True)

            qr_filename = f"{nombre}_{codigo_barras or 'sin_codigo'}_{int(time.time())}.png"
            qr_path = f"qr/{qr_filename}"

            qr_img.save(os.path.join(qr_folder, qr_filename))

            with get_db_connection() as conn:
                query = """
                    INSERT INTO productos (nombre, categoria, precio, cantidad, proveedor, codigo_barras, negocio, fecha_vencimiento, qr_path)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                conn.execute(query, (
                    nombre, categoria, precio, cantidad, proveedor,
                    codigo_barras, negocio_actual, fecha_vencimiento, qr_path
                ))
                conn.commit()

            flash(f"Producto '{nombre}' registrado en {negocio_actual}.", "success")

        except Exception as e:
            flash(f"Error al registrar producto: {str(e)}", "danger")

        return redirect(url_for("inventario"))

    return render_template("registrar_producto.html", negocio_actual=negocio_actual)


# ---------------------------
# EDITAR PRODUCTO
# ---------------------------
@app.route("/editar_producto/<int:id>", methods=["GET", "POST"])
def editar_producto(id):
    if session.get("rol") != "admin":
        flash("No tienes permisos para editar productos.", "danger")
        return redirect(url_for("inventario"))

    negocio_actual = session.get("negocio")
    if not negocio_actual:
        flash("Debes seleccionar un negocio antes de editar productos.", "warning")
        return redirect(url_for("seleccionar_negocio"))

    with get_db_connection() as conn:
        producto = conn.execute("SELECT * FROM productos WHERE id=?", (id,)).fetchone()

    if not producto:
        flash("Producto no encontrado.", "danger")
        return redirect(url_for("inventario"))

    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            categoria = request.form["categoria"].strip()
            precio = float(request.form["precio"])
            cantidad = int(request.form["cantidad"])
            proveedor = request.form.get("proveedor", "").strip()
            codigo_barras = request.form.get("codigo_barras", "").strip()
            fecha_vencimiento = request.form.get("fecha_vencimiento") or None

            qr_data = codigo_barras if codigo_barras else nombre
            qr_img = qrcode.make(qr_data)
            qr_folder = os.path.join("static", "qr")
            os.makedirs(qr_folder, exist_ok=True)
            qr_filename = f"{nombre}_{codigo_barras or 'sin_codigo'}_{int(time.time())}.png"
            qr_path = f"qr/{qr_filename}"
            qr_img.save(os.path.join(qr_folder, qr_filename))

            with get_db_connection() as conn:
                conn.execute("""
                    UPDATE productos
                    SET nombre=?, categoria=?, precio=?, cantidad=?, proveedor=?, codigo_barras=?, negocio=?, fecha_vencimiento=?, qr_path=?
                    WHERE id=?
                """, (
                    nombre, categoria, precio, cantidad, proveedor,
                    codigo_barras, negocio_actual, fecha_vencimiento, qr_path, id
                ))
                conn.commit()

            flash(f"Producto '{nombre}' actualizado correctamente.", "success")

        except Exception as e:
            flash(f"Error al editar producto: {str(e)}", "danger")

        return redirect(url_for("inventario"))

    return render_template("editar_producto.html", producto=producto, negocio_actual=negocio_actual)


# ---------------------------
# ELIMINAR PRODUCTO (ADMIN)
# ---------------------------
@app.route("/eliminar_producto/<int:id>", methods=["POST"])
def eliminar_producto(id):
    if session.get("rol") != "admin":
        flash("No tienes permisos para eliminar productos.", "danger")
        return redirect(url_for("inventario"))

    try:
        with get_db_connection() as conn:
            producto = conn.execute("SELECT nombre FROM productos WHERE id=?", (id,)).fetchone()

            if not producto:
                flash("Producto no encontrado.", "warning")
                return redirect(url_for("inventario"))

            conn.execute("DELETE FROM productos WHERE id=?", (id,))
            conn.commit()

        flash(f"Producto '{producto['nombre']}' eliminado correctamente.", "success")

    except Exception as e:
        flash(f"Error al eliminar producto: {str(e)}", "danger")

    return redirect(url_for("inventario"))


# ---------------------------
# Registrar venta
# ---------------------------
@app.route("/registrar_venta", methods=["GET", "POST"])
def registrar_venta():
    conn = get_db_connection()
    productos = conn.execute("SELECT * FROM productos").fetchall()

    if request.method == "POST":
        try:
            import json
            carrito_json = request.form.get("carrito")
            carrito = json.loads(carrito_json) if carrito_json else []

            cliente = request.form.get("cliente", "Consumidor Final")
            fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            if not carrito:
                flash("El carrito está vacío, no se puede registrar la venta.", "danger")
                conn.close()
                return redirect(url_for("registrar_venta"))

            total_venta = sum(float(item["precio"]) * int(item["cantidad"]) for item in carrito)

            cur = conn.execute(
                "INSERT INTO ventas (fecha, cliente, total, usuario_id) VALUES (?, ?, ?, ?)",
                (fecha, cliente, total_venta, session.get("usuario_id") or 0)
            )
            venta_id = cur.lastrowid

            for item in carrito:
                producto_id = int(item["id"])
                cantidad = int(item["cantidad"])
                precio = float(item["precio"])

                stock_actual = conn.execute(
                    "SELECT cantidad FROM productos WHERE id=?", (producto_id,)
                ).fetchone()["cantidad"]

                if cantidad > stock_actual:
                    flash(f"Stock insuficiente para '{item['nombre']}'. Disponible: {stock_actual}", "danger")
                    conn.rollback()
                    conn.close()
                    return redirect(url_for("registrar_venta"))

                conn.execute(
                    "INSERT INTO detalle_ventas (venta_id, producto_id, cantidad, precio) VALUES (?, ?, ?, ?)",
                    (venta_id, producto_id, cantidad, precio)
                )

                conn.execute(
                    "UPDATE productos SET cantidad = cantidad - ? WHERE id = ?",
                    (cantidad, producto_id)
                )

            conn.commit()
            flash(f"Venta registrada con éxito. Total: ${total_venta:,.2f}", "success")

        except Exception as e:
            flash(f"Error al registrar la venta: {str(e)}", "danger")
        finally:
            conn.close()

        return redirect(url_for("ventas"))

    conn.close()
    return render_template("registrar_venta.html", productos=productos)


# ---------------------------
# Ver ventas
# ---------------------------
@app.route("/ventas", methods=["GET"])
def ventas():
    conn = get_db_connection()

    # Capturar filtros desde la URL (GET)
    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")
    cliente = request.args.get("cliente")

    query = """
        SELECT v.id,
               v.fecha,
               v.cliente,
               v.total,
               u.nombre AS vendedor
        FROM ventas v
        LEFT JOIN usuarios u ON v.usuario_id = u.id
        WHERE 1=1
    """
    params = []

    # Aplicar filtros dinámicos
    if fecha_inicio:
        query += " AND v.fecha >= ?"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND v.fecha <= ?"
        params.append(fecha_fin)
    if cliente:
        query += " AND v.cliente LIKE ?"
        params.append(f"%{cliente}%")

    query += " ORDER BY v.fecha DESC"

    ventas = conn.execute(query, params).fetchall()
    conn.close()

    return render_template("ventas.html",
                           ventas=ventas,
                           fecha_inicio=fecha_inicio,
                           fecha_fin=fecha_fin,
                           cliente=cliente)


# ---------------------------
# Detalle de venta
# ---------------------------
@app.route("/ventas/<int:venta_id>")
def detalle_venta(venta_id):
    conn = get_db_connection()

    venta = conn.execute("""
        SELECT v.id, v.fecha, v.cliente, v.total, u.nombre AS vendedor
        FROM ventas v
        LEFT JOIN usuarios u ON v.usuario_id = u.id
        WHERE v.id = ?
    """, (venta_id,)).fetchone()

    detalle = conn.execute("""
        SELECT d.producto_id, p.nombre AS producto, d.cantidad, d.precio
        FROM detalle_ventas d
        JOIN productos p ON d.producto_id = p.id
        WHERE d.venta_id = ?
    """, (venta_id,)).fetchall()

    conn.close()
    return render_template("detalle_venta.html", venta=venta, detalle=detalle)
# ---------------------------
# Exportar ventas a Excel
# ---------------------------
@app.route("/ventas/export/excel")
def exportar_ventas_excel():
    import pandas as pd
    conn = get_db_connection()
    ventas = conn.execute("""
        SELECT v.id, v.fecha, v.cliente, v.total, u.nombre AS vendedor
        FROM ventas v
        LEFT JOIN usuarios u ON v.usuario_id = u.id
        ORDER BY v.fecha DESC
    """).fetchall()
    conn.close()

    df = pd.DataFrame(ventas, columns=["id", "fecha", "cliente", "total", "vendedor"])

    filename = "ventas.xlsx"
    df.to_excel(filename, index=False)

    return send_file(filename, as_attachment=True)


# ---------------------------
# Exportar ventas a PDF
# ---------------------------
@app.route("/ventas/export/pdf")
def exportar_ventas_pdf():
    from fpdf import FPDF
    conn = get_db_connection()

    ventas = conn.execute("""
        SELECT v.id, v.fecha, v.cliente, v.total, u.nombre AS vendedor
        FROM ventas v
        LEFT JOIN usuarios u ON v.usuario_id = u.id
        ORDER BY v.fecha DESC
    """).fetchall()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, "Reporte de Ventas", ln=True, align="C")
    pdf.ln(10)

    for v in ventas:
        pdf.set_font("Arial", "B", 12)
        pdf.cell(200, 10, f"Venta #{v['id']} - Cliente: {v['cliente']} - Total: ${v['total']:,.2f}", ln=True)
        pdf.set_font("Arial", size=10)
        pdf.cell(200, 8, f"Fecha: {v['fecha']} | Vendedor: {v['vendedor'] if v['vendedor'] else 'N/A'}", ln=True)

        detalle = conn.execute("""
            SELECT d.cantidad, d.precio, p.nombre AS producto
            FROM detalle_ventas d
            JOIN productos p ON d.producto_id = p.id
            WHERE d.venta_id = ?
        """, (v["id"],)).fetchall()

        for d in detalle:
            subtotal = d["cantidad"] * d["precio"]
            pdf.cell(
                200,
                8,
                f"- {d['producto']} | Cant: {d['cantidad']} | Precio: ${d['precio']:,.2f} | Subtotal: ${subtotal:,.2f}",
                ln=True
            )

        pdf.ln(5)

    conn.close()

    filename = "ventas_detalle.pdf"
    pdf.output(filename)

    return send_file(filename, as_attachment=True)



#----------------------------
# Exportar inventario a Excel con filtros
#----------------------------
@app.route("/inventario/export/excel")
def exportar_inventario_excel():
    import pandas as pd
    negocio_actual = session.get("negocio")

    if not negocio_actual:
        flash("Debes seleccionar un negocio antes de exportar.", "warning")
        return redirect(url_for("seleccionar_negocio"))

    # Capturar filtros
    categoria = request.args.get("categoria")
    stock_bajo = request.args.get("stock_bajo")
    proximos_vencer = request.args.get("proximos_vencer")

    conn = get_db_connection()
    query = """
        SELECT nombre, categoria, precio, cantidad, proveedor, codigo_barras, fecha_vencimiento
        FROM productos WHERE negocio = ?
    """
    params = [negocio_actual]

    if categoria:
        query += " AND categoria LIKE ?"
        params.append(f"%{categoria}%")

    productos = conn.execute(query, params).fetchall()
    conn.close()

    # Aplicar filtros en memoria
    current_date = datetime.now().date()
    if stock_bajo:
        productos = [p for p in productos if p["cantidad"] < 5]
    if proximos_vencer:
        productos = [
            p for p in productos
            if p["fecha_vencimiento"] and
               (datetime.strptime(p["fecha_vencimiento"], "%Y-%m-%d").date() - current_date).days < 30
        ]

    # Convertir a DataFrame con las 7 columnas seleccionadas
    df = pd.DataFrame(productos, columns=[
        "nombre","categoria","precio","cantidad","proveedor","codigo_barras","fecha_vencimiento"
    ])
    filename = "inventario_filtrado.xlsx"
    df.to_excel(filename, index=False)

    return send_file(filename, as_attachment=True)


#----------------------------
# Exportar inventario a PDF con filtros
#----------------------------
@app.route("/inventario/export/pdf")
def exportar_inventario_pdf():
    negocio_actual = session.get("negocio")

    if not negocio_actual:
        flash("Debes seleccionar un negocio antes de exportar.", "warning")
        return redirect(url_for("seleccionar_negocio"))

    # Capturar filtros
    categoria = request.args.get("categoria")
    stock_bajo = request.args.get("stock_bajo")
    proximos_vencer = request.args.get("proximos_vencer")

    conn = get_db_connection()
    query = """
        SELECT nombre, categoria, precio, cantidad, proveedor, codigo_barras, fecha_vencimiento
        FROM productos WHERE negocio = ?
    """
    params = [negocio_actual]

    if categoria:
        query += " AND categoria LIKE ?"
        params.append(f"%{categoria}%")

    productos = conn.execute(query, params).fetchall()
    conn.close()

    # Aplicar filtros en memoria
    current_date = datetime.now().date()
    if stock_bajo:
        productos = [p for p in productos if p["cantidad"] < 5]
    if proximos_vencer:
        productos = [
            p for p in productos
            if p["fecha_vencimiento"] and
               (datetime.strptime(p["fecha_vencimiento"], "%Y-%m-%d").date() - current_date).days < 30
        ]

    # Crear PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, f"Reporte de Inventario - {negocio_actual.capitalize()}", ln=True, align="C")

    # Encabezados
    pdf.set_font("Arial", "B", 12)
    pdf.cell(200, 10, "Productos filtrados", ln=True)

    # Filas
    pdf.set_font("Arial", size=10)
    for p in productos:
        pdf.cell(
            200, 8,
            f"{p['nombre']} | Cat: {p['categoria']} | Precio: ${p['precio']:,.0f} | Stock: {p['cantidad']} | Prov: {p['proveedor']} | Código: {p['codigo_barras']} | Venc: {p['fecha_vencimiento']}",
            ln=True
        )

    # Guardar archivo
    filename = f"inventario_{negocio_actual}.pdf"
    ruta = os.path.join("static", "reportes", filename)
    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    pdf.output(ruta)

    return send_file(ruta, as_attachment=True)


# ---------------------------
# Reporte
# ---------------------------
@app.route("/reporte", methods=["GET", "POST"])
def reporte():
    conn = get_db_connection()

    fecha_inicio = request.form.get("fecha_inicio")
    fecha_fin = request.form.get("fecha_fin")

    query = """
        SELECT v.cliente, v.fecha, dv.cantidad, dv.precio, p.nombre
        FROM detalle_ventas dv
        JOIN productos p ON dv.producto_id = p.id
        JOIN ventas v ON dv.venta_id = v.id
    """
    params = []
    if fecha_inicio and fecha_fin:
        query += " WHERE v.fecha BETWEEN ? AND ?"
        params = [fecha_inicio, fecha_fin]

    ventas = conn.execute(query, params).fetchall()
    conn.close()

    ganancias = sum(v["cantidad"] * v["precio"] for v in ventas)

    productos_vendidos = {}
    precios_productos = {}
    clientes = {}

    for v in ventas:
        productos_vendidos[v["nombre"]] = productos_vendidos.get(v["nombre"], 0) + v["cantidad"]
        precios_productos[v["nombre"]] = v["precio"]
        clientes[v["cliente"]] = clientes.get(v["cliente"], 0) + 1

    productos_top = sorted(productos_vendidos.items(), key=lambda x: x[1], reverse=True)
    precios_lista = [precios_productos[p[0]] for p in productos_top]
    clientes_top = sorted(clientes.items(), key=lambda x: x[1], reverse=True)

    return render_template(
        "reporte.html",
        ganancias=ganancias,
        productos_top=productos_top,
        precios_productos=precios_lista,
        clientes_top=clientes_top,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        negocio_actual=session.get("negocio", "general")
    )


# ---------------------------
# MAIN
# ---------------------------
if __name__ == "__main__":
    app.run(debug=True)